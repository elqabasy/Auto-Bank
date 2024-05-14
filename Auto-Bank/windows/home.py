# External
import datetime
import io
import os
import random
import tempfile

import pyperclip
import xlsxwriter

# My Packages
from db import DATABASE

# Me
from openpyxl import Workbook
from design.py.home import Ui_MainWindowHome

# PySide6
from settings.settings import SC, Settings
from PySide6 import QtCore, QtGui, QtNetwork, QtWebEngineCore, QtWidgets
from utils.package import NumberScrapping, ShowOtherWindow, TakeScreenShot
from utils.validation.validator import AccountNumberValidator, ArabicValidator

from .profile_info import ProfileInfo


# Home
class Home(QtWidgets.QMainWindow, Ui_MainWindowHome):
    Message = QtWidgets.QMessageBox
    timer = QtCore.QTimer()
    _url = "https://digital.banquemisr.com/bmonlinebusiness/customer-login"

    def __init__(self):
        super(Home, self).__init__()
        self.networkManager = QtNetwork.QNetworkAccessManager(self)
        self.networkManager.finished.connect(self.onFinished)
        self._profile = QtWebEngineCore.QWebEngineProfile("PROFILE_NAME")
        self.setupUi(self)
        self.setWindowFlag(QtCore.Qt.WindowType.WindowCloseButtonHint, True)
        self.initView()
        self.initBlur()
        self.initDefaults()
        self.initShortcuts()
        self.initValidators()  # high priority
        self.initEntryInputs()
        self.initButtonsActions()
        self.initWebEngineSettings()
        self.setFocus()

    def initValidators(self):
        self.lineEditCreatorName.setValidator(ArabicValidator())
        self.lineEditDefaultCustomerName.setValidator(ArabicValidator())
        self.lineEditDefaultAccountNumber.setValidator(
            AccountNumberValidator()
        )
        self.lineEditAccountCreatorAccountNumber.setValidator(
                AccountNumberValidator()
        )

        # lineEditLang

    def initWebEngineSettings(self):
        self.widgetWebPage.hide()
        self.widgetWebPage.loadStarted.connect(self.on_load_started)
        self.widgetWebPage.loadProgress.connect(self.on_load_progress)
        self.widgetWebPage.loadFinished.connect(self.on_load_finished)
        self.widgetWebPage.setPage(
            QtWebEngineCore.QWebEnginePage(
                self._profile
            )
        )
        self._profile.downloadRequested.connect(self.handle_download)

        # THEN load page
        self.widgetWebPage.load(QtCore.QUrl(self._url))

    def handle_download(self, download_item: QtCore.QEvent):
        self.Message.information(
            self,
            "خطأ", "تم تنزيل الملف في هذا المسار",
            self.Message.StandardButton.Ok
        )
        try:
            os.startfile(download_item)
            download_item.accept()
        except Exception as Error:
            self.Message.warning(
                    self,
                    "Error",
                    Error,
                    self.Message.StandardButton.Ok
                )

    def on_load_started(self):
        self.widgetWebPage.hide()
        self.labelWebsiteLoadPrpgress.show()
        self.labelWebsiteLoadPrpgress.setText("يتم تحميل الموقع الان")

    def on_load_progress(self, progress):
        # 1 2 3 .. 100
        self.labelWebsiteLoadPrpgress.setText(f"جاري التحميل ... {progress}%")

    def on_load_finished(self, ok):
        if ok:
            self.labelWebsiteLoadPrpgress.setText("تم التحميل بنجاح")
            self.widgetWebPage.show()
            self.labelWebsiteLoadPrpgress.hide()
            # be = QtWidgets.QGraphicsBlurEffect(self)
            # be.setBlurRadius(8)
            # self.widgetWebPage.setGraphicsEffect(be)
        else:
            self.labelWebsiteLoadPrpgress.setText(
                "فشل الاتحميل، اضغط علي تحديث لاعادة تحميل الصفحه"
            )

    def auto_login(self):
        if self.widgetWebPage.page().loadFinished:
            defaultUsername: str = self.lineEditDefaultUsername.text()
            defaultPassword: str = self.lineEditDefaultPassword.text()
            js_code = f"""
            var usernameField = document.getElementById('username');
            var passwordField = document.getElementById('password');

            usernameField.value = '{defaultUsername}';
            passwordField.value = '{defaultPassword}';
            """
            self.widgetWebPage.page().runJavaScript(js_code)
        else:
            self.Message.critical(
                    self,
                    "Error - Auto-Login",
                    "اعد تحميل الصفجه مجددا"
                )

    def initView(self):
        self.statusBarLabel = QtWidgets.QLabel("مستعد")
        self.statusBarLabel.setFont(QtGui.QFont("Cairo", 10))
        self.statusBarLabel.setMargin(15)
        self.statusbar.setLayoutDirection(
                QtCore.Qt.LayoutDirection.LeftToRight
            )
        self.statusbar.addPermanentWidget(self.statusBarLabel)

        self.statusBarLabel.setText("جاري تهيئة البرنامج")
        self.timer.timeout.connect(lambda: self.checkWebsiteStatus())
        self.timer.timeout.connect(
            lambda: self.statusBarLabel.setText("مستعد")
        )

        self.timer.start(3500)

        self.spinBoxTransactionAmount.textChanged.connect(
            lambda: self.labelMoney.setText(
                NumberScrapping.numberToEgyptionCurrency(
                    self.spinBoxTransactionAmount.value()
                )
            )
        )

    def Add(self):
        bankName = str(self.comboBoxBankName.currentText()).strip()
        transitionAmount = int(self.spinBoxTransactionAmount.value())
        bankNumber = str(self.comboBoxBankName.currentData()).strip()
        creatorName = str(self.lineEditCreatorName.text()).strip().title()
        creatorAccountNumber = str(
            self.lineEditAccountCreatorAccountNumber.text()
        ).strip()

        row = self.tableWidgetData.rowCount()

        # check
        if creatorName.strip() != "" and len(creatorName) >= 10:
            if creatorAccountNumber.strip() != "" and len(
                creatorAccountNumber
            ) == 16:
                if bankNumber.strip() != "" and len(bankNumber) == 11:
                    self.tableWidgetData.insertRow(row)
                    self.tableWidgetData.setItem(
                        row, 0, QtWidgets.QTableWidgetItem(creatorName)
                    )
                    self.tableWidgetData.setItem(
                        row, 1, QtWidgets.QTableWidgetItem(
                            creatorAccountNumber
                        )
                    )
                    self.tableWidgetData.setItem(
                        row, 2, QtWidgets.QTableWidgetItem(
                            str(transitionAmount)
                        )
                    )
                    self.tableWidgetData.setItem(
                        row, 3, QtWidgets.QTableWidgetItem(bankName)
                    )
                    self.tableWidgetData.setItem(
                        row, 4, QtWidgets.QTableWidgetItem(bankNumber)
                    )
                else:
                    self.Message.warning(
                        self,
                        "خطأ",
                        "اسم البنك غير صحيح، تأكد من انك اخترت اسم البنك"
                        " الصحيح، وأيضا كود البنك يتكون من 11 حرف",
                        self.Message.StandardButton.Ok,
                    )
            else:
                self.Message.warning(
                    self,
                    "خطأ",
                    "رقم الحساب غير صحيح، تأكد من انه يتكون من 16 رقم",
                    self.Message.StandardButton.Ok,
                )
        else:
            self.Message.warning(
                self,
                "خطأ",
                "اسم العميل يجب ان يكون صحيح",
                self.Message.StandardButton.Ok,
            )

    def initBlur(self):
        for child_widget in self.groupBoxAutomationParts.findChildren(
            QtWidgets.QWidget
        ):
            child_widget: QtWidgets.QWidget = child_widget
            if (
                child_widget.objectName() != "labelWarningAboutQuickParts"
                and child_widget.objectName() != "groupBoxGetReports"
                and child_widget.objectName() != "labelGetReports"
                and child_widget.objectName() != "dateEditReportFrom"
                and child_widget.objectName() != "dateEditReportTo"
                and child_widget.objectName() != "pushButtonGetDatedReport"
            ):
                be = QtWidgets.QGraphicsBlurEffect(self)
                be.setBlurRadius(2.5)
                child_widget.setGraphicsEffect(be)

    def initDefaults(self):
        # defaults
        # username, password
        username = str(SC.get(Settings.Defaults.Session.Username.value, ""))
        self.lineEditDefaultUsername.setText(username)

        password = str(SC.get(Settings.Defaults.Session.Password.value, ""))
        self.lineEditDefaultPassword.setText(password)

        defaultAccountNumber = str(
            SC.get(
                Settings.Defaults.Transaction.CreditorAccountNumber.value, ""
            )
        )
        self.lineEditDefaultAccountNumber.setText(defaultAccountNumber)

        defaultBankCode = str(
            SC.get(Settings.Defaults.Transaction.CreditorBankName.value, "")
        )
        self.lineEditDefaultBankName.setText(defaultBankCode)

        defaultCustomerName = str(
            SC.get(Settings.Defaults.Transaction.CreditorName.value, "")
        )
        self.lineEditDefaultCustomerName.setText(defaultCustomerName)

    def checkWebsiteStatus(self):
        request = QtNetwork.QNetworkRequest(QtCore.QUrl(self._url))
        self.networkManager.get(request)

    def onFinished(self, reply: QtNetwork.QNetworkReply):
        if (
            reply.attribute(
                QtNetwork.QNetworkRequest.Attribute.HttpStatusCodeAttribute
            ) == 200
        ):
            self.radioButtonAccountStatus.setChecked(True)
            self.radioButtonAccountStatus.setText("متصل")
        else:
            self.radioButtonAccountStatus.setChecked(False)
            self.radioButtonAccountStatus.setText("غير متصل")

    def initEntryInputs(self):
        self.comboBoxBankName.clear()
        comboboxesBanksData = DATABASE.getData()
        for x in comboboxesBanksData:
            name = str(x[0]).strip().title()
            serial = str(x[1]).strip().upper()
            self.comboBoxBankName.addItem(name, userData=serial)

    def initButtonsActions(self):
        self.pushButtonRefresh.clicked.connect(self.refreshApplication)
        self.pushButtonNextLevel.clicked.connect(self.openWebEngineTab)

        # copy
        self.pushButtonCopyUsername.clicked.connect(
            lambda: pyperclip.copy(self.lineEditDefaultUsername.text())
        )
        self.pushButtonCopyPassword.clicked.connect(
            lambda: pyperclip.copy(self.lineEditDefaultPassword.text())
        )
        self.pushButtonCopyBankName.clicked.connect(
            lambda: pyperclip.copy(self.lineEditDefaultBankName.text())
        )
        self.pushButtonCopyCustomerName.clicked.connect(
            lambda: pyperclip.copy(self.lineEditDefaultCustomerName.text())
        )
        self.pushButtonCopyAccountNumber.clicked.connect(
            lambda: pyperclip.copy(self.lineEditDefaultAccountNumber.text())
        )

        # files
        self.pushButtonCopyPath.clicked.connect(
            lambda: pyperclip.copy(self.lineEditXLSXPath.text())
        )

        # functionality
        self.pushButtonAdd.clicked.connect(self.Add)
        self.pushButtonDelete.clicked.connect(self.Delete)

        # defaults
        self.lineEditDefaultUsername.textChanged.connect(
            lambda: SC.set(
                Settings.Defaults.Session.Username.value,
                self.lineEditDefaultUsername.text().strip(),
            )
        )
        self.lineEditDefaultPassword.textChanged.connect(
            lambda: SC.set(
                Settings.Defaults.Session.Password.value,
                self.lineEditDefaultPassword.text().strip(),
            )
        )
        self.lineEditDefaultBankName.textChanged.connect(
            lambda: SC.set(
                Settings.Defaults.Transaction.CreditorBankName.value,
                self.lineEditDefaultBankName.text().strip(),
            )
        )
        self.lineEditDefaultCustomerName.textChanged.connect(
            lambda: SC.set(
                Settings.Defaults.Transaction.CreditorName.value,
                self.lineEditDefaultCustomerName.text().strip(),
            )
        )
        self.lineEditDefaultAccountNumber.textChanged.connect(
            lambda: SC.set(
                Settings.Defaults.Transaction.CreditorAccountNumber.value,
                self.lineEditDefaultAccountNumber.text().strip(),
            )
        )

        # radio buttons
        self.radioButtonActivateAutoLogin.clicked.connect(self.auto_login)

    def initShortcuts(self):
        # tabWidgetHome - navigation
        QtGui.QShortcut("Ctrl+Tab", self.tabWidgetHome).activated.connect(
            lambda: self.tabWidgetHome.setCurrentIndex(self.getNextTabIndex(1))
        )
        QtGui.QShortcut(
            "Ctrl+Shift+Tab",
            self.tabWidgetHome
        ).activated.connect(
            lambda: self.tabWidgetHome.setCurrentIndex(
                self.getNextTabIndex(-1)
            )
        )

        # tabData - transactions
        QtGui.QShortcut("Insert", self.tabData).activated.connect(self.Add)
        QtGui.QShortcut(
            QtGui.QKeySequence.StandardKey.Delete, self.tabData
        ).activated.connect(self.Delete)
        QtGui.QShortcut(
            QtGui.QKeySequence.StandardKey.Refresh, self.tabData
        ).activated.connect(self.refreshApplication)
        QtGui.QShortcut("F6", self.tabData).activated.connect(
            self.autoTransaction
        )

        # application - all
        QtGui.QShortcut("F1", self).activated.connect(
            lambda: ShowOtherWindow(self, ProfileInfo)
        )
        QtGui.QShortcut(
            QtGui.QKeySequence.StandardKey.Refresh,
            self
        ).activated.connect(
            lambda: self.refreshApplication()
        )
        QtGui.QShortcut("F5", self).activated.connect(
            lambda: self.refreshApplication()
        )
        QtGui.QShortcut(
            QtGui.QKeySequence.StandardKey.Print,
            self
        ).activated.connect(
            lambda: TakeScreenShot(self)
        )

    def getNextTabIndex(self, direction=1):
        """
        when:1
            return the next tab index
        else:
            return the previus tab index
        """
        tabsCount: int = self.tabWidgetHome.count()
        currentIndex: int = self.tabWidgetHome.currentIndex()
        if direction == 1:
            if (tabsCount - 1) == currentIndex:
                return 0
            else:
                return currentIndex + 1
        else:
            if currentIndex == 0:
                return tabsCount - 1
            else:
                return currentIndex - 1

    def refreshApplication(self):
        self.lineEditCreatorName.clear()
        self.tableWidgetData.setRowCount(0)
        self.widgetWebPage.load(QtCore.QUrl(self._url))
        self.spinBoxTransactionAmount.setValue(0)
        self.lineEditAccountCreatorAccountNumber.clear()

    def openWebEngineTab(self):
        self.tabWidgetHome.setCurrentWidget(self.tabUploadData)
        self.Generate_V2()

    def Delete(self):
        row = self.tableWidgetData.currentRow()
        self.tableWidgetData.removeRow(row)

    def validateLastRow(self):
        lastRowIndex = self.tableWidgetData.rowCount() - 1
        if lastRowIndex >= 0:
            customerName = self.tableWidgetData.item(lastRowIndex, 0)
            if not customerName or not customerName.text().strip():
                self.Message.warning(
                    self, "خطأ", "اسم العميل غير صحيح",
                    self.Message.StandardButton.Ok
                )
                self.tableWidgetData.setCurrentCell(lastRowIndex, 0)
                return False

            accountNumber = self.tableWidgetData.item(lastRowIndex, 1)
            if (
                not accountNumber
                or not accountNumber.text().strip()
                or len(accountNumber.text()) != 16
            ):
                self.Message.warning(
                    self,
                    "خطأ",
                    "رقم الحساب غير صحيح، تأكد من أنه يتكون من 16 رقم",
                    self.Message.StandardButton.Ok,
                )
                self.tableWidgetData.setCurrentCell(lastRowIndex, 1)
                return False

            money = self.tableWidgetData.item(lastRowIndex, 2)
            if not money or not money.text().strip():
                self.Message.warning(
                    self,
                    "خطأ",
                    "المبلغ غير صحيح، حاول ادخاله مرة اخري",
                    self.Message.StandardButton.Ok,
                )
                self.tableWidgetData.setCurrentCell(lastRowIndex, 2)
                return False

            bankCode = self.tableWidgetData.item(lastRowIndex, 3)
            if (
                not bankCode
                or not bankCode.text().strip()
                or len(bankCode.text()) != 11
            ):
                self.Message.warning(
                    self,
                    "خطأ",
                    "كود البنك غير صحيح، تأكد من انه يتكون من 11 رقم",
                    self.Message.StandardButton.Ok,
                )
                self.tableWidgetData.setCurrentCell(lastRowIndex + 1, 3)
                return False
            return True
        return True

    def Generate_V2(self):
        # Packages
        self.tableWidgetData.selectAll()
        SELECTED = [
            x for x in self.tableWidgetData.selectionModel().selectedRows()
        ]
        if len(SELECTED) > 0:
            # Create an in-memory workbook
            output = io.BytesIO()
            workbook = xlsxwriter.Workbook(output)
            Sheet1 = workbook.add_worksheet("Sheet1")

            headerData = [
                "NationalID",
                "CreditorName",
                "CreditorAccountNumber",
                "CreditorBank",
                "CreditorBankBranch",
                "TransactionAmount",
                "Comments",
                "Email",
                "Mobile",
            ]
            for index, value in enumerate(headerData):
                Sheet1.write(0, index, value)

            # set data
            ROWS = int(self.tableWidgetData.rowCount()) - 1
            fileRow = 1
            for SELECTED_Index in SELECTED:
                if SELECTED_Index.row() <= ROWS:
                    creditorNameColumn = 0
                    creditorAccountNumberColumn = 1
                    creditorBankColumn = 4
                    transactionAmountColumn = 2

                    creditorName = (
                        self.tableWidgetData.model()
                        .index(SELECTED_Index.row(), creditorNameColumn)
                        .data()
                    )
                    creditorAccountNumber = (
                        self.tableWidgetData.model()
                        .index(
                            SELECTED_Index.row(),
                            creditorAccountNumberColumn
                        )
                        .data()
                    )
                    creditorBank = (
                        self.tableWidgetData.model()
                        .index(SELECTED_Index.row(), creditorBankColumn)
                        .data()
                    )
                    transactionAmount = (
                        self.tableWidgetData.model()
                        .index(SELECTED_Index.row(), transactionAmountColumn)
                        .data()
                    )

                    Sheet1.write(fileRow, 2 - 1, creditorName)
                    Sheet1.write(fileRow, 4 - 1, creditorBank)
                    Sheet1.write(fileRow, 3 - 1, creditorAccountNumber)
                    Sheet1.write(fileRow, 6 - 1, int(transactionAmount))
                    Sheet1.write(fileRow, 7 - 1, "comment")
                    fileRow += 1

            # Close the workbook
            workbook.close()

            # Get the in-memory XLSX data
            xlsx_data = output.getvalue()

            # Create a temporary file that will be deleted after use
            with tempfile.NamedTemporaryFile(
                delete=False,
                suffix=".xlsx"
            ) as tmpfile:
                tmpfile.write(xlsx_data)
                tmpfile.flush()

                # Get the file path of the temporary file
                temp_file_path = tmpfile.name
                self.lineEditXLSXPath.setText(str(temp_file_path))
            self.Message.information(
                self,
                "نجح",
                "تم انشاء ملف الاكسل بنجاح، لن تحتاج الي فتحه وحفظه،"
                " انسخ مساره من الجزئ الايسر من تبويبة مرحلة رفع البيانات",
            )
        else:
            self.tabWidgetHome.setCurrentWidget(self.tabData)
            self.Message.warning(
                self,
                "خطأ",
                "لم يتم اضافة اي معاملات، ادخل معاملة واحدة علي الاقل",
                self.Message.StandardButton.Ok,
            )

    def Generate(self):
        self.tableWidgetData.selectAll()
        SELECTED = [
            x for x in self.tableWidgetData.selectionModel().selectedRows()
        ]
        if len(SELECTED) > 0:
            TYPE = "XLSX"
            if TYPE == "XLSX":
                # create XLSX file
                WorkBook = Workbook()
                # WorkBook.create_sheet("Sheet1")
                Sheet1 = WorkBook.active
                Sheet1.title = "Sheet1"
                # header data
                headerData = [
                    "NationalID",
                    "CreditorName",
                    "CreditorAccountNumber",
                    "CreditorBank",
                    "CreditorBankBranch",
                    "TransactionAmount",
                    "Comments",
                    "Email",
                    "Mobile",
                ]
                for index, value in enumerate(headerData):
                    Sheet1.cell(row=1, column=index + 1).value = value

                # set data
                ROWS = int(self.tableWidgetData.rowCount()) - 1
                fileRow = 1
                for SELECTED_Index in SELECTED:
                    if SELECTED_Index.row() <= ROWS:
                        creditorNameColumn = 0
                        creditorAccountNumberColumn = 1
                        creditorBankColumn = 3
                        transactionAmountColumn = 2
                        # commentsColumn = 4

                        creditorName = (
                            self.tableWidgetData.model()
                            .index(SELECTED_Index.row(), creditorNameColumn)
                            .data()
                        )
                        creditorAccountNumber = (
                            self.tableWidgetData.model()
                            .index(
                                SELECTED_Index.row(),
                                creditorAccountNumberColumn
                            )
                            .data()
                        )
                        creditorBank = (
                            self.tableWidgetData.model()
                            .index(SELECTED_Index.row(), creditorBankColumn)
                            .data()
                        )
                        transactionAmount = (
                            self.tableWidgetData.model()
                            .index(
                                SELECTED_Index.row(),
                                transactionAmountColumn
                            )
                            .data()
                        )

                        # creditorName
                        Sheet1.cell(fileRow + 1, 2).value = creditorName
                        Sheet1.cell(fileRow + 1, 4).value = creditorBank
                        Sheet1.cell(
                            fileRow + 1,
                            3
                        ).value = creditorAccountNumber
                        Sheet1.cell(
                            fileRow + 1,
                            6
                        ).value = int(transactionAmount)
                        Sheet1.cell(
                            fileRow + 1,
                            7
                        ).value = "comment"
                        fileRow += 1

                filePath = os.path.abspath(
                    "C:/Users/{user_login}/Desktop/NEW_PAYMENT_{date_time}"
                    "_time_{file_size}_size_.xlsx".format(
                        user_login=os.getlogin(),
                        file_size=random.randint(100, 1000),
                        date_time=datetime.datetime.today().strftime(
                            "%d-%m-%Y %H_%M_%S"
                        ),
                    )
                )
                WorkBook.save(filename=filePath)
                self.lineEditXLSXPath.setText(filePath)
                pyperclip.copy(self.lineEditDefaultUsername.text())
                if os.path.isfile(filePath):
                    os.startfile(filePath)
        else:
            self.tabWidgetHome.setCurrentWidget(self.tabData)
            self.Message.warning(
                self,
                "خطأ",
                "لم يتم اضافة اي معاملات، ادخل معاملة علي الاقل",
                self.Message.StandardButton.Ok,
            )

    def autoTransaction(self):
        self.lineEditCreatorName.setText(
            self.lineEditDefaultCustomerName.text()
        )
        self.lineEditAccountCreatorAccountNumber.setText(
            self.lineEditDefaultAccountNumber.text()
        )
        self.comboBoxBankName.setCurrentIndex(
            self.comboBoxBankName.findText(self.lineEditDefaultBankName.text())
        )
        self.spinBoxTransactionAmount.setValue(10)
        self.pushButtonAdd.setFocus()

    def closeEvent(self, event: QtCore.QEvent):
        event.accept()

    def keyPressEvent(self, event):
        if event.key() == QtCore.Qt.Key.Key_F9:
            os.startfile(self._url)
        if event.key() == QtCore.Qt.Key.Key_F5:
            self.refreshApplication()
        if event.key() == QtCore.Qt.Key.Key_F11:
            if self.isMaximized():
                self.showNormal()
            else:
                self.showMaximized()
