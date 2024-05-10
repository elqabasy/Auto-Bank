# External
from openpyxl import Workbook
from fake_useragent import UserAgent
import os, io, tempfile, xlsxwriter, pyperclip, os, random, datetime

# My Packages
from source.database import DATABASE
from source.settings import Settings
from source.profile_info import ProfileInfo
from source.package import NumberScrapping, ShowOtherWindow, TakeScreenShot


# PySide6
from PySide6.QtCore import QUrl, QTimer, Qt
from PySide6.QtGui import QShortcut, QKeySequence, QFont
from PySide6.QtWebEngineCore import QWebEngineProfile, QWebEnginePage
from PySide6.QtNetwork import QNetworkAccessManager, QNetworkRequest, QNetworkReply
from PySide6.QtWidgets import QMainWindow, QMessageBox, QTableWidgetItem, QGraphicsBlurEffect, QWidget, QLabel, QApplication


from design.py.home import Ui_MainWindowHome
class Home(QMainWindow, Ui_MainWindowHome):
    Message = QMessageBox
    timer = QTimer()
    _settings = Settings("Mahros AL-Qabasy", "Auto-Bank")
    _url = "https://digital.banquemisr.com/bmonlinebusiness/customer-login"
    def __init__(self):
        super(Home,self).__init__()
        self.networkManager = QNetworkAccessManager(self)
        self.networkManager.finished.connect(self.onFinished)
        self._profile = QWebEngineProfile("PROFILE_NAME")
        self.setupUi(self)
        self.setWindowFlag(Qt.WindowType.WindowCloseButtonHint, True)

        self.initView()
        self.initDeafaults()
        self.initShortcuts()
        self.initEntryInputs()
        self.initButtonsActions()
        self.initWebEngineSettings()
        self.initBlureStyle()
        # self.Generate_V2()

    def initWebEngineSettings(self):
        self.widgetWebPage.hide()
        self.widgetWebPage.loadStarted.connect(self.on_load_started)
        self.widgetWebPage.loadProgress.connect(self.on_load_progress)
        self.widgetWebPage.loadFinished.connect(self.on_load_finished)


        # get random user agent
        ua = UserAgent()
        random_user_agent = ua.random
        self._profile.setHttpUserAgent(random_user_agent)
        self.widgetWebPage.setPage(QWebEnginePage(self._profile)) # (QWebEngineProfile(self._profile).defaultPage())
        self._profile.downloadRequested.connect(self.handle_download)
        # self.widgetWebPage.


        # self.widgetWebPage.settings().setAttribute(QWebEngineSettings.WebAttribute., False)

        # THEN load page
        self.widgetWebPage.load(QUrl(self._url))

    def handle_download(self, download_item):
        self.showSuccess("تم تنزيل الملف في هذا المسار")
        try:
            os.startfile(download_item)
            download_item.accept()
        except Exception as Error:
            self.showError(Error)

    def on_load_started(self):
        self.widgetWebPage.hide()
        self.labelWebsiteLoadPrpgress.show()
        self.labelWebsiteLoadPrpgress.setText("يتم تحميل الموقع الان")

    def on_load_progress(self, progress):
        # 1 2 3 .. 100
        # print(progress)
        self.labelWebsiteLoadPrpgress.setText(f"جاري التحميل ... {progress}%")

    def on_load_finished(self, ok):
        if ok:
            self.labelWebsiteLoadPrpgress.setText("تم التحميل بنجاح")
            self.widgetWebPage.show()
            self.labelWebsiteLoadPrpgress.hide()
            # be = QGraphicsBlurEffect(self)
            # be.setBlurRadius(8)
            # self.widgetWebPage.setGraphicsEffect(be)
        else:
            self.labelWebsiteLoadPrpgress.setText("فشل الاتحميل، اضغط علي تحديث لاعادة تحميل الصفحه")


    def auto_login(self):
        # print(self.widgetWebPage.page().loadFinished)
        if self.widgetWebPage.page().loadFinished:
            defaultUsername:str = self.lineEditUsername.text()
            defaultPassword:str = self.lineEditPassword.text()
            js_code = f"""
            var usernameField = document.getElementById('username');
            var passwordField = document.getElementById('password');

            usernameField.value = '{defaultUsername}';
            passwordField.value = '{defaultPassword}';
            """
            # var loginButton = document.getElementById('loginButton');
            # loginButton.click();
            self.widgetWebPage.page().runJavaScript(js_code)
        else:
            self.Message.critical(self, "Error - Auto-Login", "اعد تحميل الصفجه مجددا")
            

    def initView(self):
        self.statusBarLabel = QLabel("مستعد")
        self.statusBarLabel.setFont(QFont("Cairo", 10))
        self.statusBarLabel.setMargin(15)
        self.statusbar.setLayoutDirection(Qt.LayoutDirection.LeftToRight)
        self.statusbar.addPermanentWidget(self.statusBarLabel)


        self.statusBarLabel.setText("جاري تهيئة البرنامج")
        self.timer.timeout.connect(lambda: self.checkWebsiteStatus())
        self.timer.timeout.connect(lambda: self.statusBarLabel.setText("مستعد"))
        
        self.timer.start(3500)

        self.spinBoxTransactionAmount.textChanged.connect(lambda: self.labelMoney.setText(NumberScrapping.numberToEgyptionCurrency(self.spinBoxTransactionAmount.value())))

    def Add(self):
        bankName = str(self.comboBoxBankName.currentText()).strip()
        transitionAmount = int(self.spinBoxTransactionAmount.value())
        bankNumber = str(self.comboBoxBankNumber.currentText()).strip()
        creatorName = str(self.lineEditCreatorName.text()).strip().title()
        creatorAccountNumber = str(self.lineEditAccountCreatorAccountNumber.text()).strip()

        row = self.tableWidgetData.rowCount()

        # check
        if creatorName.strip() != "":
            if creatorAccountNumber.strip() != "" and len(creatorAccountNumber) == 16:
                if bankNumber.strip() != "" and len(bankNumber) == 11:
                    # if package
                    # rowCount = self.tableWidgetData.rowCount()
                    self.tableWidgetData.insertRow(row)
                    self.tableWidgetData.setItem(row, 0, QTableWidgetItem(creatorName))
                    self.tableWidgetData.setItem(row, 1, QTableWidgetItem(creatorAccountNumber))
                    self.tableWidgetData.setItem(row, 2, QTableWidgetItem(str(transitionAmount)))
                    self.tableWidgetData.setItem(row, 3, QTableWidgetItem(bankName))
                    self.tableWidgetData.setItem(row, 4, QTableWidgetItem(bankNumber))
                else:
                    self.showError("اسم البنك غير صحيح، تأكد من انك اخترت اسم البنك الصحيح، وأيضا كود البنك يتكون من 11 حرف")
            else:
                self.showError("رقم الحساب غير صحيح، تأكد من انه يتكون من 16 رقم")
        else:
            self.showError("اسم العميل يجب ان يكون صحيح")


    def initBlureStyle(self):
        for child_widget in self.groupBoxAutomationParts.findChildren(QWidget):
            child_widget:QWidget = child_widget
            if child_widget.objectName() != "labelWarningAboutQuickParts" and child_widget.objectName() != "groupBoxGetReports" and child_widget.objectName() != "labelGetReports" and child_widget.objectName() != "dateEditReportFrom" and child_widget.objectName() != "dateEditReportTo" and child_widget.objectName() != "pushButtonGetDatedReport":
                be = QGraphicsBlurEffect(self)
                be.setBlurRadius(2.5)
                child_widget.setGraphicsEffect(be)



    def initDeafaults(self):
        # defaults
        # username, password
        username = str(self._settings.load_setting("data/defaultUsername", ""))
        self.lineEditUsername.setText(username)
        self.lineEditUsername.setText(username)

        password = str(self._settings.load_setting("data/defaultPassword", ""))
        self.lineEditPassword.setText(password)

        defaultAccountNumber = str(self._settings.load_setting("data/defaultAccountNumber", ""))
        self.lineEditDefaultAccountNumber.setText(defaultAccountNumber)
        
        defaultBankCode = str(self._settings.load_setting("data/defaultBankCode", ""))
        self.lineEditDefaultBankCode.setText(defaultBankCode)
        
        defaultCustomerName = str(self._settings.load_setting("data/defaultCustomerName", ""))
        self.lineEditDefaultCustomerName.setText(defaultCustomerName)

    def checkWebsiteStatus(self):
        request = QNetworkRequest(QUrl(self._url))
        self.networkManager.get(request)

    def onFinished(self, reply:QNetworkReply):
        if reply.attribute(QNetworkRequest.Attribute.HttpStatusCodeAttribute) == 200:
            self.radioButtonAccountStatus.setChecked(True)
            self.radioButtonAccountStatus.setText("متصل")
        else:
            self.radioButtonAccountStatus.setChecked(False)
            self.radioButtonAccountStatus.setText("غير متصل")

    def initEntryInputs(self):
        self.comboBoxBankName.clear()
        self.comboBoxBankNumber.clear()
        comboboxesBanksData = DATABASE.Query("SELECT Name, Serial FROM `Banks`;")
        for x in comboboxesBanksData:
            name = x[0]
            serial = x[1]
            self.comboBoxBankName.addItem(str(name).strip().title())
            self.comboBoxBankNumber.addItem(str(serial).strip())
        

        # Banks
        def getBankCode():
            name = str(self.comboBoxBankName.currentText()).strip().title()
            if name:
                comboboxesAccountsData = DATABASE.Query("SELECT Serial FROM `Banks` WHERE Name = ?", (name,))
                self.comboBoxBankNumber.clear()
                for x in comboboxesAccountsData:
                    number = x[0]
                    self.comboBoxBankNumber.addItem(str(number).strip())
                    self.comboBoxBankNumber.setCurrentIndex(0)

        self.comboBoxBankName.currentTextChanged.connect(getBankCode)

    def initButtonsActions(self):
        self.pushButtonRefresh.clicked.connect(self.refreshApplication)
        self.pushButtonNextLevel.clicked.connect(self.openWebEngineTab)
        # self.pushButtonAddNewTransaction.clicked.connect(self.addNewPayment)


        # copy
        self.pushButtonCopyUsername.clicked.connect(lambda: pyperclip.copy(self.lineEditUsername.text()))
        self.pushButtonCopyPassword.clicked.connect(lambda: pyperclip.copy(self.lineEditPassword.text()))
        self.pushButtonCopyBankCode.clicked.connect(lambda: pyperclip.copy(self.lineEditDefaultBankCode.text()))
        self.pushButtonCopyCustomerName.clicked.connect(lambda: pyperclip.copy(self.lineEditDefaultCustomerName.text()))
        self.pushButtonCopyAccountNumber.clicked.connect(lambda: pyperclip.copy(self.lineEditDefaultAccountNumber.text()))
        
        # files
        self.pushButtonCopyPath.clicked.connect(lambda: pyperclip.copy(self.lineEditXLSXPath.text()))

        # functionality
        self.pushButtonAdd.clicked.connect(self.Add)
        self.pushButtonDelete.clicked.connect(self.Delete)

        
        # self.lineEditUsername.connect(lambda: pyperclip.copy(self.lineEditXLSXPath.text()))
        # self.pushButtonCopyUsername.clicked.connect(lambda: pyperclip.copy(self.lineEditUsername.text()))
        # defaults
        self.lineEditUsername.textChanged.connect(lambda: self._settings.save_setting("data/defaultUsername", self.lineEditUsername.text().strip()))
        self.lineEditPassword.textChanged.connect(lambda: self._settings.save_setting("data/defaultPassword", self.lineEditPassword.text().strip()))
        self.lineEditDefaultCustomerName.textChanged.connect(lambda: self._settings.save_setting("data/defaultCustomerName", self.lineEditDefaultCustomerName.text().strip()))
        self.lineEditDefaultBankCode.textChanged.connect(lambda: self._settings.save_setting("data/defaultBankCode", self.lineEditDefaultBankCode.text().strip()))
        self.lineEditDefaultAccountNumber.textChanged.connect(lambda: self._settings.save_setting("data/defaultAccountNumber", self.lineEditDefaultAccountNumber.text().strip()))

        # radio buttons
        self.radioButtonActivateAutoLogin.clicked.connect(self.auto_login)

    def initShortcuts(self):
        # tabWidgetHome - navigation
        bavigateTabsRight = QShortcut("Ctrl+Tab", self.tabWidgetHome).activated.connect(lambda: self.tabWidgetHome.setCurrentIndex(self.getNextTabIndex(1)))
        bavigateTabsLeft = QShortcut("Ctrl+Shift+Tab", self.tabWidgetHome).activated.connect(lambda: self.tabWidgetHome.setCurrentIndex(self.getNextTabIndex(-1)))

        # tabData - transactions
        addTransaction = QShortcut("Insert", self.tabData).activated.connect(self.Add)
        deleteTransaction = QShortcut(QKeySequence.StandardKey.Delete, self.tabData).activated.connect(self.Delete)
        clearTransactions = QShortcut(QKeySequence.StandardKey.Refresh, self.tabData).activated.connect(self.refreshApplication)
        autoTransaction = QShortcut("F6", self.tabData).activated.connect(self.autoTransaction)

        # application - all 
        openProfileInfo = QShortcut("F1", self).activated.connect(lambda: ShowOtherWindow(self, ProfileInfo))
        refreshApplication = QShortcut(QKeySequence.StandardKey.Refresh, self).activated.connect(lambda: self.tabWidgetHome.setCurrentIndex(self.getNextTabIndex(-1)))
        takeScreenShot = QShortcut(QKeySequence.StandardKey.Print, self).activated.connect(lambda: TakeScreenShot(self))
        # if event.key() == Qt.Key.Key_Insert:
        #     self.Add()
        # if event.key() == Qt.Key.Key_Delete:
        #     self.Delete()

        # if event.key() == Qt.Key.Key_F1:
        #     self.profile_info_ = ProfileInfo()
        #     self.profile_info_.show()
        # if event.key() == Qt.Key.Key_F5:
        #     self.initView()
        # if event.key() == Qt.Key.Key_F6:
        #     self.lineEditCreatorName.setText(self.lineEditDefaultCustomerName.text())
        #     self.lineEditAccountCreatorAccountNumber.setText(self.lineEditDefaultAccountNumber.text())
        #     self.spinBoxTransactionAmount.setValue(10)
        #     self.comboBoxBankName.setFocus()
        #     self.comboBoxBankName.showPopup()
        # if event.key() == Qt.Key.Key_F9:
        #     os.startfile(self._url)
        # if event.key() == Qt.Key.Key_F11:
        #     if self.isMaximized():
        #         self.showNormal()
        #     else:
        #         self.showMaximized()

        # addNewPayment = QShortcut("Ctrl+Shift+I",self).activated.connect(self.Screenshot)
        # clearWindow = QShortcut("Ctrl+Shift+C",self).activated.connect(self.ClearLoggings)
        # refreshWindow = QShortcut(QKeySequence.StandardKey.Refresh,self).activated.connect(self.initView)


    def getNextTabIndex(self, direction=1):
        """
        when:1 
            return the next tab index 
        else:
            return the previus tab index
        """
        tabsCount:int = self.tabWidgetHome.count()
        currentIndex:int = self.tabWidgetHome.currentIndex()
        if direction == 1:
            if (tabsCount-1) == currentIndex:
                return 0
            else:
                return currentIndex+1
        else:
            if currentIndex == 0:
                return tabsCount-1
            else:
                return currentIndex-1
        
    def refreshApplication(self):
        self.tableWidgetData.setRowCount(0)
        self.widgetWebPage.load(QUrl(self._url))
        # self.widgetWebPag
        
    def closeCurrentSession(self):
        pass
    
    def openWebEngineTab(self):
        self.tabWidgetHome.setCurrentWidget(self.tabUploadData)
        # self.save_table_to_excel()
        # self.Generate()
        self.Generate_V2()
    
    def Delete(self):
        row = self.tableWidgetData.currentRow()
        self.tableWidgetData.removeRow(row)

    def addNewPayment(self):
        # if self.validateLastRow():
        #     rowCount = self.tableWidgetData.rowCount()
        #     self.tableWidgetData.insertRow(rowCount)
        #     self.tableWidgetData.setCurrentCell(rowCount, 0)
        #     self.tableWidgetData.setItem(0, 0, QTableWidgetItem(self.lineEditDefaultCustomerName.text()))
        #     self.tableWidgetData.setItem(0, 1, QTableWidgetItem(self.lineEditDefaultAccountNumber.text()))
        #     self.tableWidgetData.setItem(0, 2, QTableWidgetItem("10"))
        #     self.tableWidgetData.setItem(0, 3, QTableWidgetItem(self.lineEditDefaultBankCode.text()))
        pass

    def validateLastRow(self):
        lastRowIndex = self.tableWidgetData.rowCount() - 1
        if lastRowIndex >= 0:
            customerName = self.tableWidgetData.item(lastRowIndex, 0)
            if not customerName or not customerName.text().strip():
                self.showError("اسم العميل غير صحيح")
                self.tableWidgetData.setCurrentCell(lastRowIndex, 0)
                return False
            
            accountNumber = self.tableWidgetData.item(lastRowIndex, 1)
            if not accountNumber or not accountNumber.text().strip() or len(accountNumber.text()) != 16:
                self.showError("رقم الحساب غير صحيح، تأكد من أنه يتكون من 16 رقم")
                self.tableWidgetData.setCurrentCell(lastRowIndex, 1)
                return False

            money = self.tableWidgetData.item(lastRowIndex, 2)
            if not money or not money.text().strip():
                self.showError("المبلغ غير صحيح، حاول ادخاله مرة اخري")
                self.tableWidgetData.setCurrentCell(lastRowIndex, 2)
                return False
            
            bankCode = self.tableWidgetData.item(lastRowIndex, 3)
            if not bankCode or not bankCode.text().strip() or len(bankCode.text()) != 11:
                self.showError("كود البنك غير صحيح، تأكد من انه يتكون من 11 رقم")
                self.tableWidgetData.setCurrentCell(lastRowIndex+1, 3)
                return False
            return True
        return True

    def Generate_V2(self):
        # pckages
        
        self.tableWidgetData.selectAll()
        SELECTED = [x for x in self.tableWidgetData.selectionModel().selectedRows()]
        if len(SELECTED) > 0:
            # Create an in-memory workbook
            output = io.BytesIO()
            workbook = xlsxwriter.Workbook(output)
            Sheet1 = workbook.add_worksheet("Sheet1")

            headerData = ["NationalID", "CreditorName", "CreditorAccountNumber", "CreditorBank", "CreditorBankBranch", "TransactionAmount", "Comments", "Email", "Mobile"]
            for index, value in enumerate(headerData):
                # Sheet1.cell(row=1, column=index+1).value = value
                Sheet1.write(0, index, value)

            # set data
            ROWS = int(self.tableWidgetData.rowCount()) - 1
            fileRow = 1
            for SELECTED_Index in SELECTED:
                if SELECTED_Index.row() <= ROWS:
                    creditorNameColumn =  0
                    creditorAccountNumberColumn =  1
                    creditorBankColumn =  4
                    transactionAmountColumn =  2
            #         commentsColumn =  4

                    creditorName = self.tableWidgetData.model().index(SELECTED_Index.row(), creditorNameColumn).data()
                    creditorAccountNumber = self.tableWidgetData.model().index(SELECTED_Index.row(), creditorAccountNumberColumn).data()
                    creditorBank = self.tableWidgetData.model().index(SELECTED_Index.row(), creditorBankColumn).data()
                    transactionAmount = self.tableWidgetData.model().index(SELECTED_Index.row(), transactionAmountColumn).data()
                    # comments = self.tableWidgetData.model().index(SELECTED_Index.row(), commentsColumn).data()

            #         # creditorName
                    Sheet1.write(fileRow, 2-1, creditorName)
                    Sheet1.write(fileRow, 4-1, creditorBank)
                    Sheet1.write(fileRow, 3-1, creditorAccountNumber)
                    Sheet1.write(fileRow, 6-1, int(transactionAmount))
                    Sheet1.write(fileRow, 7-1, "comment")
                    fileRow += 1

            # Close the workbook
            workbook.close()

            # Get the in-memory XLSX data
            xlsx_data = output.getvalue()
            
            # Create a temporary file that will be deleted after use
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmpfile:
                tmpfile.write(xlsx_data)
                tmpfile.flush()

                # Get the file path of the temporary file
                temp_file_path = tmpfile.name
                self.lineEditXLSXPath.setText(str(temp_file_path))
                # print(temp_file_path)
                # os.startfile(temp_file_path)
        else:
            self.tabWidgetHome.setCurrentWidget(self.tabData)
            self.showError("لم يتم اضافة اي معاملات، ادخل معاملة واحدة علي الاقل")

    def Generate(self):
        self.tableWidgetData.selectAll()
        SELECTED = [x for x in self.tableWidgetData.selectionModel().selectedRows()]
        if len(SELECTED) > 0:
            TYPE = "XLSX"
            if TYPE == "XLSX":
                # create XLSX file
                WorkBook = Workbook()
                # WorkBook.create_sheet("Sheet1")
                Sheet1 = WorkBook.active
                Sheet1.title = "Sheet1"
                # header data
                headerData = ["NationalID", "CreditorName", "CreditorAccountNumber", "CreditorBank", "CreditorBankBranch", "TransactionAmount", "Comments", "Email", "Mobile"]
                for index, value in enumerate(headerData):
                    Sheet1.cell(row=1, column=index+1).value = value

                # set data
                ROWS = int(self.tableWidgetData.rowCount()) - 1
                fileRow = 1
                for SELECTED_Index in SELECTED:
                    if SELECTED_Index.row() <= ROWS:
                        creditorNameColumn =  0
                        creditorAccountNumberColumn =  1
                        creditorBankColumn =  3
                        transactionAmountColumn =  2
                        commentsColumn =  4

                        creditorName = self.tableWidgetData.model().index(SELECTED_Index.row(), creditorNameColumn).data()
                        creditorAccountNumber = self.tableWidgetData.model().index(SELECTED_Index.row(), creditorAccountNumberColumn).data()
                        creditorBank = self.tableWidgetData.model().index(SELECTED_Index.row(), creditorBankColumn).data()
                        transactionAmount = self.tableWidgetData.model().index(SELECTED_Index.row(), transactionAmountColumn).data()
                        # comments = self.tableWidgetData.model().index(SELECTED_Index.row(), commentsColumn).data()

                        # creditorName
                        Sheet1.cell(fileRow+1, 2).value = creditorName
                        Sheet1.cell(fileRow+1, 4).value = creditorBank
                        Sheet1.cell(fileRow+1, 3).value = creditorAccountNumber
                        Sheet1.cell(fileRow+1, 6).value = int(transactionAmount)
                        Sheet1.cell(fileRow+1, 7).value = "comment"
                        fileRow += 1

                filePath = os.path.abspath("C:/Users/{user_login}/Desktop/NEW_PAYMENT_{date_time}_time_{file_size}_size_.xlsx".format(user_login=os.getlogin(), file_size=random.randint(100, 1000), date_time=datetime.datetime.today().strftime("%d-%m-%Y %H_%M_%S")))
                WorkBook.save(filename=filePath)
                self.lineEditXLSXPath.setText(filePath)
                pyperclip.copy(self.lineEditUsername.text())
                if os.path.isfile(filePath):
                    os.startfile(filePath)
        else:
            self.tabWidgetHome.setCurrentWidget(self.tabData)
            self.showError("لم يتم اضافة اي معاملات، ادخل معاملة علي الاقل")

    def showError(self, message):
        QMessageBox.critical(self, "خطأ", message)

    def showSuccess(self, message):
        QMessageBox.information(self, "نجاح", message)

    def closeEvent(self, event):
        pass
    
    def autoTransaction(self):
        self.lineEditCreatorName.setText(self.lineEditDefaultCustomerName.text())
        self.lineEditAccountCreatorAccountNumber.setText(self.lineEditDefaultAccountNumber.text())
        self.spinBoxTransactionAmount.setValue(10)
        self.comboBoxBankName.setFocus()
        self.comboBoxBankName.showPopup()

    # def showErrorMessage(self, text:str, title:str="Error", informative_text:str=""):
    #     Message = QMessageBox(QMessageBox.Icon.Critical, title, text, QMessageBox.StandardButton.Ok | QMessageBox.StandardButton.Yes)
    #     Message.setInformativeText(informative_text)
    #     Message.exec_()
    #     if Message == QMessageBox.StandardButton.Yes:
    #         pyperclip.copy(str())

    def keyPressEvent(self, event): 
        if event.key() == Qt.Key.Key_F9:
            os.startfile(self._url)
        if event.key() == Qt.Key.Key_F11:
            if self.isMaximized():
                self.showNormal()
            else:
                self.showMaximized()