import package
from PySide6 import QtCore
from openpyxl import Workbook
from database import DATABASE
from settings import Settings
from PySide6.QtCore import QUrl
import pyperclip, os, random, datetime
from design.py.home import Ui_MainWindowHome
from PySide6.QtWidgets import QMainWindow, QMessageBox, QTableWidgetItem
from PySide6.QtNetwork import QNetworkAccessManager, QNetworkRequest, QNetworkReply
class Home(QMainWindow, Ui_MainWindowHome):
    Message = QMessageBox
    timer = QtCore.QTimer()
    _settings = Settings("Mahros AL-Qabasy", "Auto-Bank")
    _url = "https://digital.banquemisr.com/bmonlinebusiness/customer-login"
    def __init__(self):
        super(Home,self).__init__()
        self.networkManager = QNetworkAccessManager(self)
        self.networkManager.finished.connect(self.onFinished)
        self.setupUi(self)
        self.setWindowFlag(QtCore.Qt.WindowType.WindowCloseButtonHint, True)
        self.widgetWebPage.hide()

        self.initView()
        self.initDeafaults()
        self.initEntryInputs()
        self.initButtonsActions()

        self.widgetWebPage.loadStarted.connect(self.on_load_started)
        self.widgetWebPage.loadProgress.connect(self.on_load_progress)
        self.widgetWebPage.loadFinished.connect(self.on_load_finished)

        self.widgetWebPage.load(QUrl(self._url))


    def on_load_started(self):
        self.labelWebsiteLoadPrpgress.setText("يتم تحميل الموقع الان")
    def on_load_progress(self, progress):
        self.labelWebsiteLoadPrpgress.setText(f"جاري التحميل ... {progress}%")
    def on_load_finished(self, ok):
        if ok:
            self.labelWebsiteLoadPrpgress.setText("تم التحميل بنجاح")
            self.labelWebsiteLoadPrpgress.hide()
            self.widgetWebPage.show()
        else:
            self.labelWebsiteLoadPrpgress.setText("فشل الاتحميل")

    def initView(self):
        self.statusbar.showMessage("جاري تهيئة البرنامج")
        self.timer.timeout.connect(lambda: self.checkWebsiteStatus())
        self.timer.timeout.connect(lambda: self.statusbar.showMessage("مستعد"))
        
        self.timer.start(3500)

        self.spinBoxTransactionAmount.textChanged.connect(lambda: self.labelMoney.setText(package.NumberScrapping.numberToEgyptionCurrency(self.spinBoxTransactionAmount.value())))

    def Add(self):
        comments = "comment"
        transitionAmount = int(self.spinBoxTransactionAmount.value())
        bankNumber = str(self.comboBoxBankNumber.currentText()).strip()
        creatorName = str(self.lineEditCreatorName.text()).strip().title()
        creatorAccountNumber = str(self.lineEditAccountCreatorAccountNumber.text()).strip()

        row = self.tableWidgetData.rowCount()

        # check
        if creatorName.strip() != "":
            if creatorAccountNumber.strip() != "" and len(creatorAccountNumber) == 16:
                if bankNumber.strip() != "" and len(bankNumber) == 11:
                    pass
                    # rowCount = self.tableWidgetData.rowCount()
                    self.tableWidgetData.insertRow(row)
                    self.tableWidgetData.setItem(row, 0, QTableWidgetItem(creatorName))
                    self.tableWidgetData.setItem(row, 1, QTableWidgetItem(creatorAccountNumber))
                    self.tableWidgetData.setItem(row, 2, QTableWidgetItem(str(transitionAmount)))
                    self.tableWidgetData.setItem(row, 3, QTableWidgetItem(bankNumber))
                else:
                    self.showError("اسم البنك غير صحيح، تأكد من انك اخترت اسم البنك الصحيح، وأيضا كود البنك يتكون من 11 حرف")
            else:
                self.showError("رقم الحساب غير صحيح، تأكد من انه يتكون من 16 رقم")
        else:
            self.showError("اسم العميل يجب ان يكون صحيح")

               
    def initDeafaults(self):
        # defaults
        # username, password
        username = str(self._settings.load_setting("data/defaultUsername", ""))
        self.lineEditUsername.setText(username)
        self.lineEditUsername.setText(username)

        password = str(self._settings.load_setting("data/defaultPassword", ""))
        self.lineEditPassword.setText(password)

        defaultAccountNumber = str(self._settings.load_setting("data/defaultAccountNumber", ""))
        self.lineEditDefaultAccountNumber.setText(defaultAccountNumber)
        
        defaultBankCode = str(self._settings.load_setting("data/defaultBankCode", ""))
        self.lineEditDefaultBankCode.setText(defaultBankCode)
        
        defaultCustomerName = str(self._settings.load_setting("data/defaultCustomerName", ""))
        self.lineEditDefaultCustomerName.setText(defaultCustomerName)

    def checkWebsiteStatus(self):
        request = QNetworkRequest(QUrl(self._url))
        self.networkManager.get(request)

    def onFinished(self, reply:QNetworkReply):
        if reply.attribute(QNetworkRequest.Attribute.HttpStatusCodeAttribute) == 200:
            self.radioButtonAccountStatus.setChecked(True)
            self.radioButtonAccountStatus.setText("متصل")
        else:
            self.radioButtonAccountStatus.setChecked(False)
            self.radioButtonAccountStatus.setText("غير متصل")

    def initEntryInputs(self):
        self.comboBoxBankName.clear()
        self.comboBoxBankNumber.clear()
        comboboxesBanksData = DATABASE.Query("SELECT Name, Serial FROM `Banks`;")
        for x in comboboxesBanksData:
            name = x[0]
            serial = x[1]
            self.comboBoxBankName.addItem(str(name).strip().title())
            self.comboBoxBankNumber.addItem(str(serial).strip())
        

        # Banks
        def getBankCode():
            name = str(self.comboBoxBankName.currentText()).strip().title()
            if name:
                comboboxesAccountsData = DATABASE.Query("SELECT Serial FROM `Banks` WHERE Name = ?", (name,))
                self.comboBoxBankNumber.clear()
                for x in comboboxesAccountsData:
                    number = x[0]
                    self.comboBoxBankNumber.addItem(str(number).strip())
                    self.comboBoxBankNumber.setCurrentIndex(0)

        self.comboBoxBankName.currentTextChanged.connect(getBankCode)

    def initButtonsActions(self):
        self.pushButtonRefresh.clicked.connect(self.refreshApplication)
        self.pushButtonNextLevel.clicked.connect(self.openWebEngineTab)
        # self.pushButtonAddNewTransaction.clicked.connect(self.addNewPayment)


        # copy
        self.pushButtonCopyUsername.clicked.connect(lambda: pyperclip.copy(self.lineEditUsername.text()))
        self.pushButtonCopyPassword.clicked.connect(lambda: pyperclip.copy(self.lineEditPassword.text()))
        self.pushButtonCopyBankCode.clicked.connect(lambda: pyperclip.copy(self.lineEditDefaultBankCode.text()))
        self.pushButtonCopyCustomerName.clicked.connect(lambda: pyperclip.copy(self.lineEditDefaultCustomerName.text()))
        self.pushButtonCopyAccountNumber.clicked.connect(lambda: pyperclip.copy(self.lineEditDefaultAccountNumber.text()))
        
        # files
        self.pushButtonCopyPath.clicked.connect(lambda: pyperclip.copy(self.lineEditXLSXPath.text()))

        # functionality
        self.pushButtonAdd.clicked.connect(self.Add)
        self.pushButtonDelete.clicked.connect(self.Delete)

        
        # self.lineEditUsername.connect(lambda: pyperclip.copy(self.lineEditXLSXPath.text()))
        # self.pushButtonCopyUsername.clicked.connect(lambda: pyperclip.copy(self.lineEditUsername.text()))
        # defaults
        self.lineEditUsername.textChanged.connect(lambda: self._settings.save_setting("data/defaultUsername", self.lineEditUsername.text().strip()))
        self.lineEditPassword.textChanged.connect(lambda: self._settings.save_setting("data/defaultPassword", self.lineEditPassword.text().strip()))
        self.lineEditDefaultCustomerName.textChanged.connect(lambda: self._settings.save_setting("data/defaultCustomerName", self.lineEditDefaultCustomerName.text().strip()))
        self.lineEditDefaultBankCode.textChanged.connect(lambda: self._settings.save_setting("data/defaultBankCode", self.lineEditDefaultBankCode.text().strip()))
        self.lineEditDefaultAccountNumber.textChanged.connect(lambda: self._settings.save_setting("data/defaultAccountNumber", self.lineEditDefaultAccountNumber.text().strip()))


        # animation
        # icon = QIcon(QIcon.fromTheme(u"edit-paste"))
        # self.pushButtonCopyPath.clicked.connect(lambda: self.pushButtonCopyPath.setText("تم النسخ"))
        # self.pushButtonCopyPath.clicked.connect(lambda: self.pushButtonCopyPath.setIcon(icon))
        
        # self.pushButtonCopyUsername.clicked.connect(lambda: self.pushButtonCopyUsername.setText("تم النسخ"))
        # self.pushButtonCopyUsername.clicked.connect(lambda: self.pushButtonCopyUsername.setIcon(icon))

    def initShortcuts(self):
        # addNewPayment = QShortcut("Ctrl+Shift+I",self).activated.connect(self.Screenshot)
        # addNewPayment = QShortcut("Ctrl+Shift+I",self).activated.connect(self.Screenshot)
        # clearWindow = QShortcut("Ctrl+Shift+C",self).activated.connect(self.ClearLoggings)
        # refreshWindow = QShortcut(QKeySequence.StandardKey.Refresh,self).activated.connect(self.initView)
        pass

    # Functionality
    def refreshApplication(self):
        self.tableWidgetData.setRowCount(0)
        self.widgetWebPage.load(QUrl(self._url))
        # self.widgetWebPag
        
    def closeCurrentSession(self):
        pass
    
    def openWebEngineTab(self):
        self.tabWidgetHome.setCurrentWidget(self.tabUploadData)
        # self.save_table_to_excel()
        self.Generate()
    
    def Delete(self):
        row = self.tableWidgetData.currentRow()
        self.tableWidgetData.removeRow(row)

    def addNewPayment(self):
        # if self.validateLastRow():
        #     rowCount = self.tableWidgetData.rowCount()
        #     self.tableWidgetData.insertRow(rowCount)
        #     self.tableWidgetData.setCurrentCell(rowCount, 0)
        #     self.tableWidgetData.setItem(0, 0, QTableWidgetItem(self.lineEditDefaultCustomerName.text()))
        #     self.tableWidgetData.setItem(0, 1, QTableWidgetItem(self.lineEditDefaultAccountNumber.text()))
        #     self.tableWidgetData.setItem(0, 2, QTableWidgetItem("10"))
        #     self.tableWidgetData.setItem(0, 3, QTableWidgetItem(self.lineEditDefaultBankCode.text()))
        pass

    def validateLastRow(self):
        lastRowIndex = self.tableWidgetData.rowCount() - 1
        if lastRowIndex >= 0:
            customerName = self.tableWidgetData.item(lastRowIndex, 0)
            if not customerName or not customerName.text().strip():
                self.showError("اسم العميل غير صحيح")
                self.tableWidgetData.setCurrentCell(lastRowIndex, 0)
                return False
            
            accountNumber = self.tableWidgetData.item(lastRowIndex, 1)
            if not accountNumber or not accountNumber.text().strip() or len(accountNumber.text()) != 16:
                self.showError("رقم الحساب غير صحيح، تأكد من أنه يتكون من 16 رقم")
                self.tableWidgetData.setCurrentCell(lastRowIndex, 1)
                return False

            money = self.tableWidgetData.item(lastRowIndex, 2)
            if not money or not money.text().strip():
                self.showError("المبلغ غير صحيح، حاول ادخاله مرة اخري")
                self.tableWidgetData.setCurrentCell(lastRowIndex, 2)
                return False
            
            bankCode = self.tableWidgetData.item(lastRowIndex, 3)
            if not bankCode or not bankCode.text().strip() or len(bankCode.text()) != 11:
                self.showError("كود البنك غير صحيح، تأكد من انه يتكون من 11 رقم")
                self.tableWidgetData.setCurrentCell(lastRowIndex+1, 3)
                return False
            return True
        return True

    def Generate(self):
        self.tableWidgetData.selectAll()
        SELECTED = [x for x in self.tableWidgetData.selectionModel().selectedRows()]
        if len(SELECTED) > 0:
            TYPE = "XLSX"
            if TYPE == "XLSX":
                # create XLSX file
                WorkBook = Workbook()
                # WorkBook.create_sheet("Sheet1")
                Sheet1 = WorkBook.active
                Sheet1.title = "Sheet1"
                # header data
                # headerData = ["CreditorName", "CreditorAccountNumber", "CreditorBank", "TransactionAmount", "Comments"]
                # headerData = ["NationalID", "CreditorName", "CreditorAccountNumber", "CreditorBank", "TransactionAmount", "Comments"]
                headerData = ["NationalID", "CreditorName", "CreditorAccountNumber", "CreditorBank", "CreditorBankBranch", "TransactionAmount", "Comments", "Email", "Mobile"]
                for index, value in enumerate(headerData):
                    Sheet1.cell(row=1, column=index+1).value = value

                # set data
                ROWS = int(self.tableWidgetData.rowCount()) - 1
                fileRow = 1
                for SELECTED_Index in SELECTED:
                    if SELECTED_Index.row() <= ROWS:
                        creditorNameColumn =  0
                        creditorAccountNumberColumn =  1
                        creditorBankColumn =  3
                        transactionAmountColumn =  2
                        commentsColumn =  4

                        creditorName = self.tableWidgetData.model().index(SELECTED_Index.row(), creditorNameColumn).data()
                        creditorAccountNumber = self.tableWidgetData.model().index(SELECTED_Index.row(), creditorAccountNumberColumn).data()
                        creditorBank = self.tableWidgetData.model().index(SELECTED_Index.row(), creditorBankColumn).data()
                        transactionAmount = self.tableWidgetData.model().index(SELECTED_Index.row(), transactionAmountColumn).data()
                        # comments = self.tableWidgetData.model().index(SELECTED_Index.row(), commentsColumn).data()

                        # creditorName
                        Sheet1.cell(fileRow+1, 2).value = creditorName
                        Sheet1.cell(fileRow+1, 4).value = creditorBank
                        Sheet1.cell(fileRow+1, 3).value = creditorAccountNumber
                        Sheet1.cell(fileRow+1, 6).value = int(transactionAmount)
                        Sheet1.cell(fileRow+1, 7).value = "comment"
                        fileRow += 1

                filePath = os.path.abspath("C:/Users/{user_login}/Desktop/NEW_PAYMENT_{date_time}_time_{file_size}_size_.xlsx".format(user_login=os.getlogin(), file_size=random.randint(100, 1000), date_time=datetime.datetime.today().strftime("%d-%m-%Y %H_%M_%S")))
                WorkBook.save(filename=filePath)
                self.lineEditXLSXPath.setText(filePath)
                pyperclip.copy(self.lineEditUsername.text())
                if os.path.isfile(filePath):
                    os.startfile(filePath)
                    # Result = self.Message.information(self, "Success!", "{TYPE} Generated successfully.,\n path:\n '{filePath}' {TYPE}, \nDo you want to copy source?".format(TYPE=TYPE, filePath=filePath), buttons=self.Message.Yes|self.Message.No|self.Message.Ok, defaultButton=self.Message.Yes)
                    # if Result == self.Message.Yes:
                        # Result = self.Message.information(self, "Coping Source...", "Source File Copied successfully.", buttons=self.Message.Ok, defaultButton=self.Message.Ok)
        else:
            self.tabWidgetHome.setCurrentWidget(self.tabData)
            self.showError("لم يتم اضافة اي معاملات، ادخل معاملة علي الاقل")

    def showError(self, message):
        QMessageBox.critical(self, "خطأ", message)

    def showSuccess(self, message):
        QMessageBox.information(self, "نجاح", message)

    def closeEvent(self, event):
        if os.path.exists(self.lineEditXLSXPath.text()):
            os.remove(self.lineEditXLSXPath.text())
        event.accept()

    def keyPressEvent(self, event): 
        self.statusbar.showMessage("KeyEvent, {key}".format(key=event.modifiers().name)) 
        if event.key() == QtCore.Qt.Key.Key_Escape:
            pass
        if event.key() == QtCore.Qt.Key.Key_F11:
            if self.isMaximized():
                self.showNormal()
            else:
                self.showMaximized()
        if event.key() == QtCore.Qt.Key.Key_F5:
            self.initView()
        if event.key() == QtCore.Qt.Key.Key_Insert:
            self.Add()
        if event.key() == QtCore.Qt.Key.Key_Delete:
            self.Delete()
        if event.key() == QtCore.Qt.Key.Key_F9:
            os.startfile(self._url)